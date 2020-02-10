from unittest.mock import call, Mock, MagicMock
import pytest

from openpyxl import Workbook

import global_var as GL
from main import (
    run,
    execute_by_configID,
    handle_maps,
    handle_formulas
)


def test_run(mocker, mock_data_template):
    template = mock_data_template
    configs = template[GL.GL_TEMPLATE_KEY_CONFIGS_NAME]
    mock_make_template = mocker.patch('main.make_template',
                                      return_value=template)
    mock_execute_by_configID = mocker.patch('main.execute_by_configID')
    calls = [
        call(template, configs[1], 1),
        call(template, configs[2], 2)
    ]
    run()
    mock_make_template.assert_called_once()
    assert mock_execute_by_configID.call_count == 2
    mock_execute_by_configID.assert_has_calls(calls)


def test_execute_by_configID(mocker, mock_data_template):
    # mock Workbook object
    mock_wb = Mock()
    mock_ws = Mock()
    mock_wb.__getitem__ = Mock(return_value=mock_ws)
    mock_wb.save = Mock()
    mock_wb.close = Mock()
    mock_ws.insert_rows = Mock()
    # mock parameters
    template = mock_data_template
    configID = 1
    config = template[GL.GL_TEMPLATE_KEY_CONFIGS_NAME][configID]
    # mock functions
    mock_get_data_file_full_path = mocker.patch(
        'main.get_data_file_full_path',
        return_value='/test/data_file_full_name')
    mock_get_out_file_full_path = mocker.patch(
        'main.get_out_file_full_path',
        return_value='/test/out_file_full_name'
    )
    mock_load_workbook = mocker.patch(
        'main.load_workbook',
        return_value=mock_wb
    )
    mock_handle_maps = mocker.patch('main.handle_maps')
    mock_handle_formulas = mocker.patch('main.handle_formulas')
    # execute test function
    execute_by_configID(template, config, configID)
    # assert
    mock_get_data_file_full_path.assert_called_once_with('附件5A.xlsx')
    mock_get_out_file_full_path.assert_called_once_with('附件5A.xlsx')
    mock_load_workbook.assert_called_once_with('/test/data_file_full_name')
    mock_wb.__getitem__.assert_called_once()
    assert mock_ws.insert_rows.call_count == 2
    calls = [call(7), call(8)]
    mock_ws.insert_rows.assert_has_calls(calls)
    assert mock_handle_maps.call_count == 2
    assert mock_handle_formulas.call_count == 2
    mock_wb.save.assert_called_once_with('/test/out_file_full_name')
    mock_wb.close.assert_not_called()


def test_error_execute_by_configID(mocker, mock_data_template):
    # mock Workbook object
    mock_wb = Mock()
    mock_ws = Mock()
    mock_wb.__getitem__ = Mock(return_value=mock_ws)
    mock_wb.save = Mock()
    mock_wb.close = Mock()
    mock_ws.insert_rows = Mock()
    # mock parameters
    template = mock_data_template
    configID = 1
    config = template[GL.GL_TEMPLATE_KEY_CONFIGS_NAME][configID]
    # mock functions
    mock_get_data_file_full_path = mocker.patch(
        'main.get_data_file_full_path',
        return_value='/test/data_file_full_name')
    mock_get_out_file_full_path = mocker.patch(
        'main.get_out_file_full_path',
        return_value='/test/out_file_full_name'
    )
    mock_load_workbook = mocker.patch(
        'main.load_workbook',
        side_effect=Exception()
    )
    mock_handle_maps = mocker.patch('main.handle_maps')
    mock_handle_formulas = mocker.patch('main.handle_formulas')
    with pytest.raises(Exception):
        # execute test function
        execute_by_configID(template, config, configID)
        # assert
        mock_get_data_file_full_path.assert_called_once_with('附件5A.xlsx')
        mock_get_out_file_full_path.assert_called_once_with('附件5A.xlsx')
        mock_load_workbook.assert_called_once_with('/test/data_file_full_name')
        mock_wb.__getitem__.assert_not_called()
        mock_ws.insert_rows.assert_not_called()
        mock_handle_maps.call_count.assert_not_called()
        mock_handle_formulas.call_count.assert_not_called()
        mock_wb.save.assert_not_called()
        mock_wb.close.assert_called_once()


def test_handle_maps(mocker, mock_data_data, mock_data_map):
    # mock Worksheet object
    mock_ws = Mock()
    mock_ws.__setitem__ = Mock()
    # mock parameters
    mock_data = mock_data_data[1][0]
    mock_maps = mock_data_map[1]
    mock_row_number = 10
    # execute test function
    handle_maps(mock_ws, mock_data, mock_maps, mock_row_number)
    # assert
    assert mock_ws.__setitem__.call_count == 2
    calls = [
        call('B10', '小招通二期'),
        call('C10', '依据国家发改委对双创建设的要求')
    ]
    mock_ws.__setitem__.assert_has_calls(calls)
    ############### configID=2 ######################
    mock_ws.__setitem__.reset_mock()
    # mock parameters
    mock_data = mock_data_data[2][0]
    mock_maps = mock_data_map[2]
    mock_row_number = 9
    # execute test function
    handle_maps(mock_ws, mock_data, mock_maps, mock_row_number)
    # assert
    assert mock_ws.__setitem__.call_count == 3
    calls = [
        call('B9', '小招通四期'),
        call('C9', '依据国家发改委对'),
        call('D9', '项目D的内容')
    ]
    mock_ws.__setitem__.assert_has_calls(calls)


def test_handle_formulas(mocker, mock_data_formula):
    # mock Worksheet object
    mock_ws = Mock()
    mock_ws.__setitem__ = Mock()
    # mock parameters
    mock_formulas = mock_data_formula[1]
    mock_row_number = 6
    # execute test function
    handle_formulas(mock_ws, mock_formulas, mock_row_number)
    # assert
    assert mock_ws.__setitem__.call_count == 2
    calls = [
        call('L6', '=I6'),
        call('N6', '=SUM(O6:T6)')
    ]
    mock_ws.__setitem__.assert_has_calls(calls)
