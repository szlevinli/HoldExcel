import os
from pathlib import Path
from openpyxl import Workbook, load_workbook


def main():
    current_work_dir = Path(os.getcwd())
    FILE_NAME = '附件5A.xlsx'
    file_full_path = current_work_dir / 'data' / FILE_NAME
    wb = load_workbook(file_full_path)
    ws = wb.active
    print(f'B7 = {ws["B7"].value}; F7 = {ws["F7"].value}')


if __name__ == "__main__":
    main()
