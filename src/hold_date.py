import os
import datetime
from pathlib import Path
from openpyxl import Workbook, load_workbook


def main():
    current_work_dir = Path(os.getcwd())
    FILE_NAME = 'hold_date.xlsx'
    file_full_path = current_work_dir / 'out' / FILE_NAME
    wb = Workbook()
    ws = wb.active
    toady = datetime.datetime.today()
    cell_A1 = ws['A1']
    cell_A1.value = toady
    cell_A1.number_format = 'YYYY-MM-DD'
    wb.save(file_full_path)


if __name__ == "__main__":
    main()