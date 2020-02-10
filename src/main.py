from openpyxl import load_workbook

from generate_template import make_template
import global_var as GL
from tools import get_data_file_full_path, get_out_file_full_path


def run():
    """主程序
    """
    # 从 template.xlsx 文件中读取数据
    # 数据将以字典的格式返回
    template = make_template()
    # 从 template 字典对象中提取 'Configs' 数据
    configs = template[GL.GL_TEMPLATE_KEY_CONFIGS_NAME]
    # 遍历 'Configs'
    for configID in configs:
        # 执行每个 config
        execute_by_configID(template, configs[configID], configID)


def execute_by_configID(template, config, configID):
    """执行 config

    Args:
        template (Dictionary): 从 template.xlsx 文件中读取数据
        config (Dictionary): config配置信息
        configID (int): 配置ID
    """
    # 从 config 配置信息中读取数据文件的名称
    # 该数据文件是需要将数据填入的文件
    # 用于最终生成输出文件
    data_file_name = config[GL.GL_EXCEL_HEADER_FILE_NAME]
    data_file_full_name = get_data_file_full_path(data_file_name)
    # 输出文件名称
    out_file_full_name = get_out_file_full_path(data_file_name)
    # 需要操作的 sheet name
    sheet_name = config[GL.GL_EXCEL_HEADER_SHEET_NAME]
    # 数据插入的行
    insert_before_row = config[GL.GL_EXCEL_HEADER_INSERT_BEFORE_ROW]
    # 从 template 字典中提取指定 configID 的
    # 数据信息、映射关系和公式信息
    datum = template[GL.GL_TEMPLATE_KEY_DATUM_NAME][configID]
    maps = template[GL.GL_TEMPLATE_KEY_MAPS_NAME][configID]
    formulas = template[GL.GL_TEMPLATE_KEY_FORMULAS_NAME][configID]
    try:
        # 读取数据文件生成 Workbook 对象
        wb_data = load_workbook(data_file_full_name)
        # 读取指定的 sheet 生成 Worksheet 对象
        ws_data = wb_data[sheet_name]
        # 遍历数据信息
        for data in datum:
            # 在指定位置插入新行
            ws_data.insert_rows(insert_before_row)
            # 根据映射关系将数据写入指定的文件位置
            handle_maps(ws_data, data, maps, insert_before_row)
            # 将公式写入指定文件位置
            handle_formulas(ws_data, formulas, insert_before_row)
            insert_before_row += 1
        # 保存文件
        wb_data.save(out_file_full_name)
    except:
        wb_data.close()
        raise


def handle_maps(ws, data, maps, row_number):
    """根据映射关系将数据写入文件

    Args:
        ws (Worksheet): 工作表对象
        data (Dictionary): 待写入数据
        maps (Dictionary): 映射关系
        row_number (int): 写入位置的行号
    """
    # 遍历映射信息
    for _map in maps:
        # 在指定的位置写入数据
        cell_name = f'{_map[GL.GL_EXCEL_HEADER_TO]}{row_number}'
        _data = data[_map[GL.GL_EXCEL_HEADER_FROM]]
        ws[cell_name] = _data


def handle_formulas(ws, formulas, row_number):
    """公式数据写入文件

    Args:
        ws (Worksheet): openpyxl.Worksheet 对象
        formulas (Dictionary): 公式信息
        row_number (int): 写入位置的行号
    """
    # 遍历公式
    for formula in formulas:
        # 在指定的位置写入数据
        cell_name = (f'{formula[GL.GL_EXCEL_HEADER_COLUMN_NAME]}'
                     f'{row_number}')
        value = formula[GL.GL_EXCEL_HEADER_FORMULA].replace(
            GL.GL_EXCEL_CELL_CONTENT_REPLACE_ROW, str(row_number))
        ws[cell_name] = f'={value}'


if __name__ == "__main__":
    run()
