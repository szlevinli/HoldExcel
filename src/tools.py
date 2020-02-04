import os
import yaml
from pathlib import Path
from openpyxl import load_workbook
import global_var as GL


def read_config_yaml():
    """读取YAML格式配置文件

    Raises:
        FileNotFoundError: 读取配置文件错误

    Returns:
        Dictionary: 配置信息
    """
    cwd = get_current_work_dir()
    file_name = 'config.yml'
    full_file_name = cwd / file_name
    if Path.is_file(full_file_name):
        with open(full_file_name, 'r') as stream:
            yml = yaml.load(stream, Loader=yaml.FullLoader)
            return yml
    else:
        raise FileNotFoundError(full_file_name)


def get_current_work_dir():
    """当前工作目录

    Returns:
        Path: 当前工作目录
    """
    return Path(os.getcwd())


def get_template_file_full_path():
    """模板文件全路径，包含文件名

    Returns:
        Path: 模板文件全路径，包含文件名
    """
    config_yml = read_config_yaml()
    template_dir = config_yml[GL.GL_CONFIG_YAML_KEY_TEMPLATE_DIR]
    template_file_name = config_yml[GL.GL_CONFIG_YAML_KEY_TEMPLATE_FILE_NAME]
    return get_current_work_dir() / template_dir / template_file_name


def get_data_file_full_path(file_name):
    """数据文件全路径，包含文件名

    Args:
        file_name (string): 文件名称含扩展名

    Returns:
        Path: 数据文件全路径，包含文件名
    """
    data_dir = read_config_yaml()[GL.GL_CONFIG_YAML_KEY_DATA_DIR]
    return get_current_work_dir() / data_dir / file_name


def get_template():
    """返回模板文件

    Returns:
        Workbook: 模板文件
    """
    return load_workbook(get_template_file_full_path(), read_only=True)


def make_dict_repeatable(wb, sheet_name):
    """创建字典，允许原数据中的key值有重复
    组成的字典格式如下：
    {
        key: [{key: value, ...}, ...]
        ...
    }

    Args:
        wb (openpyxl.Workbook): excel文件对象
        sheet_name (string): excle工作表名词

    Returns:
        Dictionary: 字典，允许原数据中的key值有重复
    """
    ws = wb[sheet_name]
    dicts = {}
    rows = []
    for row in ws.iter_rows(min_row=ws.min_row, max_row=ws.max_row,
                            min_col=ws.min_column, max_col=ws.max_column,
                            values_only=True):
        rows.append(row)
    keys = rows[0]
    for row in rows[1:]:
        tmp1 = dicts[row[0]] if row[0] in dicts else []
        tmp2 = {}
        for k, v in enumerate(row[1:]):
            tmp2[keys[k+1]] = v
        tmp1.append(tmp2)
        dicts[row[0]] = tmp1
    return dicts


def make_dict_unrepeatable(wb, sheet_name):
    """创建字典，不允许原数据中的key值有重复
    组成的字典格式如下：
    {
        key: {key: value, ...}
        ...
    }

    Args:
        wb (openpyxl.Workbook): excel文件对象
        sheet_name (string): excle工作表名词

    Returns:
        Dictionary: 字典，不允许原数据中的key值有重复
    """
    ws = wb[sheet_name]
    rows = []
    for v in ws.iter_rows(min_row=ws.min_row, max_row=ws.max_row,
                          min_col=ws.min_column, max_col=ws.max_column,
                          values_only=True):
        rows.append(v)
    dicts = {}
    # 第一行是配置项名称
    keys = rows[0]
    # 循环配置信息，排除第一行的配置项名称
    for row in rows[1:]:
        tmp = {}
        # 循环配置项，排除第一个配置项
        # 第一个配置项是 configID，将作为 key
        for k, v in enumerate(row[1:]):
            tmp[keys[k+1]] = v
        dicts[row[0]] = tmp
    return dicts


def make_list(wb, sheet_name):
    """创建列表
    [
        {key: value, ...}
        ...
    ]

    Args:
        wb (openpyxl.Workbook): excel文件对象
        sheet_name (string): excle工作表名词

    Returns:
        Dictionary: 根据excel原数据生成的列表
    """
    ws = wb[sheet_name]
    rows = []
    for v in ws.iter_rows(min_row=ws.min_row, max_row=ws.max_row,
                          min_col=ws.min_column, max_col=ws.max_column,
                          values_only=True):
        rows.append(v)
    lists = []
    # 第一行是配置项名称
    keys = rows[0]
    # 循环配置信息，排除第一行的配置项名称
    for row in rows[1:]:
        tmp = {}
        for k, v in enumerate(row):
            tmp[keys[k]] = v
        lists.append(tmp)
    return lists


def get_config_from_template(template, configID):
    """从 template 对象中获取指定 configID 的 config 字典值

    Args:
        template (Dictionary): template 对象
        configID (string): 配置项ID

    Returns:
        Dictionary: config 字典值
    """
    return template[GL.GL_TEMPLATE_KEY_CONFIGS_NAME][configID]


def get_data_file_name_from_template(template, configID):
    """从 template 对象中获取指定 configID 的数据文件名称

    Args:
        template (Dictionary): template 对象
        configID (string): 配置项ID

    Returns:
        string: 数据文件名称
    """
    config = get_config_from_template(template, configID)
    data_file_name = config[GL.GL_EXCEL_HEADER_FILE_NAME]
    return data_file_name


if __name__ == "__main__":
    print(f'config.yaml: {read_config_yaml()}')
